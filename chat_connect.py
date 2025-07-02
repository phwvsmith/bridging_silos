import random
from typing import List, Set
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

def read_csv_files() -> (pd.DataFrame, pd.DataFrame):
    files = [f for f in os.listdir('.') if os.path.isfile(f) and f.endswith('.csv')]
    email_file = next((f for f in files if 'emails' in f), None)
    history_file = next((f for f in files if 'history' in f), None)
    if not email_file or not history_file:
        raise FileNotFoundError("Required files not found in the current directory.")

    emails_df = pd.read_csv(email_file)
    history_df = pd.read_csv(history_file)

    # Normalize emails: lowercase and strip whitespace
    emails_df = emails_df.applymap(lambda x: x.strip().lower() if isinstance(x, str) else x)
    history_df = history_df.applymap(lambda x: x.strip().lower() if isinstance(x, str) else x)

    return emails_df, history_df

def rename_columns(emails: pd.DataFrame, history: pd.DataFrame) -> (pd.DataFrame, pd.DataFrame):
    if len(emails.columns) != 1 or len(history.columns) != 2:
        raise ValueError("One of the DataFrames does not have the expected number of columns.")
    emails.columns = ['colleague_emails']
    history.columns = ['combination_history_1', 'combination_history_2']
    return emails, history

def dataframes_to_lists(emails_df: pd.DataFrame, history_df: pd.DataFrame) -> (list, list):
    emails_list = emails_df.iloc[:, 0].tolist()
    history_list = [set(row) for row in history_df.itertuples(index=False, name=None)]
    return emails_list, history_list

def create_new_combinations(emails_list: List[str], history_list: List[Set[str]]) -> List[Set[str]]:
    repeat_check = []
    new_combinations = []

    attempts = 0
    max_attempts = len(emails_list) * 3

    while emails_list and attempts < max_attempts:
        if len(emails_list) < 2:
            break
        email1, email2 = random.sample(emails_list, 2)
        new_set = {email1, email2}

        if new_set not in history_list and email1 not in repeat_check and email2 not in repeat_check:
            new_combinations.append(new_set)
            repeat_check.extend(new_set)
            emails_list.remove(email1)
            emails_list.remove(email2)
            attempts = 0
        else:
            attempts += 1

    df = pd.DataFrame(new_combinations, columns=['Email 1', 'Email 2'])
    filename = f"batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)

    workbook = load_workbook(filename)
    sheet = workbook.active
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 1)
        cell.font = Font(color="FF0000")
    workbook.save(filename)

    print(f"Saved to {filename}")
    return new_combinations

def update_history_and_save_csv(new_combinations: list, history_list: list):
    if new_combinations:
        history_list.extend(new_combinations)

    history_df = pd.DataFrame(history_list, columns=['combination_history_1', 'combination_history_2'])

    history_file = next((f for f in os.listdir('.') if os.path.isfile(f) and 'history' in f and f.endswith('.csv')), None)
    if not history_file:
        raise FileNotFoundError("History CSV file not found.")

    history_df.to_csv(history_file, index=False)

# Run the full process
emails_df, history_df = read_csv_files()
emails_df, history_df = rename_columns(emails_df, history_df)
emails_list, history_list = dataframes_to_lists(emails_df, history_df)
new_combinations = create_new_combinations(emails_list, history_list)
update_history_and_save_csv(new_combinations, history_list)
